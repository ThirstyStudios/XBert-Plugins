#!/usr/bin/env python3
"""Render an XBert schedule as a .xlsx workbook from a structured payload."""

from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path
from typing import Any


def _emit(result: dict[str, Any], exit_code: int) -> None:
    print(json.dumps(result), flush=True)
    sys.exit(exit_code)


def _load_payload(path: Path) -> dict[str, Any]:
    with path.open("r", encoding="utf-8") as f:
        return json.load(f)


def _build_workbook(payload: dict, out_path: Path) -> tuple[int, int]:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    wb.remove(wb.active)

    sheets = payload.get("sheets") or []
    if not sheets:
        wb.create_sheet("Sheet1")

    total_cells = 0
    number_format_default = payload.get("number_format") or "#,##0.00;(#,##0.00)"

    for sheet_def in sheets:
        name = (sheet_def.get("name") or "Sheet")[:31]
        ws = wb.create_sheet(title=name)

        columns = sheet_def.get("columns") or []
        if columns:
            ws.append(columns)
            header_row = ws[1]
            for cell in header_row:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="1C1B41", end_color="1C1B41", fill_type="solid")
                cell.alignment = Alignment(vertical="center")
            total_cells += len(columns)

        rows = sheet_def.get("rows") or []
        for row in rows:
            ws.append(list(row))
            total_cells += len(row)

        for row in ws.iter_rows(min_row=2 if columns else 1):
            for cell in row:
                if isinstance(cell.value, (int, float)):
                    cell.number_format = sheet_def.get("number_format") or number_format_default

        widths = sheet_def.get("column_widths") or []
        for idx, width in enumerate(widths, start=1):
            try:
                ws.column_dimensions[get_column_letter(idx)].width = float(width)
            except Exception:  # noqa: BLE001
                continue

        if sheet_def.get("freeze_top_row") and columns:
            ws.freeze_panes = "A2"

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out_path))
    return len(sheets) or 1, total_cells


def main() -> None:
    parser = argparse.ArgumentParser(description="Render an XBert schedule as .xlsx")
    parser.add_argument("--payload", required=True, type=Path)
    parser.add_argument("--out", required=True, type=Path)
    args = parser.parse_args()

    if not args.payload.exists():
        _emit({"status": "errors_found", "error": f"payload not found: {args.payload}"}, 2)

    payload = _load_payload(args.payload)

    try:
        sheet_count, cell_count = _build_workbook(payload, args.out)
    except ModuleNotFoundError as exc:
        _emit(
            {
                "status": "errors_found",
                "error": f"missing dependency: {exc.name}. Install: pip install openpyxl pandas",
            },
            3,
        )
    except Exception as exc:  # noqa: BLE001
        _emit({"status": "errors_found", "error": f"{type(exc).__name__}: {exc}"}, 4)

    result = {
        "path": str(args.out.resolve()),
        "exists": args.out.exists(),
        "size_bytes": args.out.stat().st_size if args.out.exists() else 0,
        "sheet_count": sheet_count,
        "cell_count": cell_count,
        "status": "ok" if (args.out.exists() and args.out.stat().st_size > 1024) else "errors_found",
    }
    _emit(result, 0 if result["status"] == "ok" else 5)


if __name__ == "__main__":
    main()
