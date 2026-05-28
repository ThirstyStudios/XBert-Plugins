#!/usr/bin/env python3
"""Recalculate an .xlsx with LibreOffice (if available) and scan for error cells.

Mirrors the contract of Anthropic's official xlsx skill recalc.py: returns a
JSON line on stdout with status "ok" or "errors_found" and a list of offending
cells.
"""

from __future__ import annotations

import json
import shutil
import subprocess
import sys
from pathlib import Path
from typing import Any

ERROR_CODES = {"#REF!", "#DIV/0!", "#VALUE!", "#N/A", "#NAME?", "#NULL!", "#NUM!"}


def _emit(result: dict[str, Any], exit_code: int) -> None:
    print(json.dumps(result), flush=True)
    sys.exit(exit_code)


def _recalc_with_libreoffice(path: Path, timeout: int) -> bool:
    """Try to force formula recalculation via LibreOffice headless. Best-effort."""
    soffice = shutil.which("soffice") or shutil.which("libreoffice")
    if not soffice:
        return False
    try:
        subprocess.run(
            [
                soffice,
                "--headless",
                "--calc",
                "--convert-to",
                "xlsx",
                "--outdir",
                str(path.parent),
                str(path),
            ],
            timeout=timeout,
            check=True,
            capture_output=True,
        )
        return True
    except Exception:  # noqa: BLE001
        return False


def _scan_for_errors(path: Path) -> list[dict[str, str]]:
    from openpyxl import load_workbook

    wb = load_workbook(str(path), data_only=False)
    offenders: list[dict[str, str]] = []
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                value = cell.value
                if isinstance(value, str) and value in ERROR_CODES:
                    offenders.append({"sheet": ws.title, "cell": cell.coordinate, "error": value})
    return offenders


def main() -> None:
    if len(sys.argv) < 2:
        _emit({"status": "errors_found", "error": "usage: recalc.py <file.xlsx> [timeout_seconds]"}, 2)

    path = Path(sys.argv[1])
    timeout = int(sys.argv[2]) if len(sys.argv) > 2 else 30

    if not path.exists():
        _emit({"status": "errors_found", "error": f"file not found: {path}"}, 2)

    recalculated = _recalc_with_libreoffice(path, timeout)

    try:
        offenders = _scan_for_errors(path)
    except ModuleNotFoundError as exc:
        _emit(
            {
                "status": "errors_found",
                "error": f"missing dependency: {exc.name}. Install: pip install openpyxl",
            },
            3,
        )

    result = {
        "path": str(path.resolve()),
        "recalculated_with_libreoffice": recalculated,
        "errors": offenders,
        "status": "ok" if not offenders else "errors_found",
    }
    _emit(result, 0 if result["status"] == "ok" else 6)


if __name__ == "__main__":
    main()
